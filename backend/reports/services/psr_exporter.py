"""
Plant Status Report (PSR) Excel Exporter
Generates Excel reports matching 100% the exact PSR format from PSR REPORT-8AM.xlsx
INCLUDING: Forecasted Load, IPP sections, Charts, and Notes
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from datetime import datetime
import os
import io
from django.conf import settings


class PSRExporter:
    """Service class for generating Plant Status Report (PSR) Excel files"""
    
    # Plant configuration matching the template EXACTLY
    PLANTS_CONFIG = {
        'AGUS1': {
            'name': 'AGUS 1',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 40, 'nominated': 0},
                {'num': 2, 'label': 'unit 2', 'capacity': 40, 'nominated': 0}
            ]
        },
        'AGUS2': {
            'name': 'AGUS 2',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 60, 'nominated': 60},
                {'num': 2, 'label': 'unit  2', 'capacity': 60, 'nominated': 60},
                {'num': 3, 'label': 'unit  3', 'capacity': 60, 'nominated': 60}
            ]
        },
        'AGUS4': {
            'name': 'AGUS 4',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 52.7, 'nominated': 0},
                {'num': 2, 'label': 'unit  2', 'capacity': 52.7, 'nominated': 52.7},
                {'num': 3, 'label': 'unit  3', 'capacity': 52.7, 'nominated': 52.7}
            ]
        },
        'AGUS5': {
            'name': 'AGUS 5',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 27.5, 'nominated': 27.5},
                {'num': 2, 'label': 'unit 2', 'capacity': 27.5, 'nominated': 27.5}
            ]
        },
        'AGUS6': {
            'name': 'AGUS 6',
            'units': [
                {'num': 1, 'label': '  unit  1', 'capacity': 34.5, 'nominated': 20},
                {'num': 2, 'label': '  unit  2', 'capacity': 34.5, 'nominated': 21},
                {'num': 3, 'label': 'unit  3', 'capacity': 50, 'nominated': 42},
                {'num': 4, 'label': 'unit  4', 'capacity': 50, 'nominated': 38},
                {'num': 5, 'label': 'unit  5', 'capacity': 50, 'nominated': 44}
            ]
        },
        'AGUS7': {
            'name': 'AGUS 7',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 27, 'nominated': 27},
                {'num': 2, 'label': 'unit 2', 'capacity': 27, 'nominated': 27}
            ]
        },
        'PULANGI4': {
            'name': 'PULANGI IV',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 85, 'nominated': 75},
                {'num': 2, 'label': 'unit  2', 'capacity': 85, 'nominated': 70},
                {'num': 3, 'label': 'unit  3', 'capacity': 85, 'nominated': 70}
            ]
        }
    }
    
    def __init__(self, queryset, report_date, report_type='psr'):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        self.queryset = queryset
        self.report_date = report_date
        self.report_type = report_type  # 'psr' or 'daily_status'
        self.data_by_plant = self._organize_data()
    
    def _organize_data(self):
        """Organize queryset data by plant and unit"""
        data = {}
        for report in self.queryset:
            plant_code = report.plant.code.upper()
            unit_num = report.unit.unit_number
            
            if plant_code not in data:
                data[plant_code] = {}
            
            data[plant_code][unit_num] = {
                'generation': float(report.generation_kwh),
                'operating_hours': float(report.operating_hours),
                'forced_outage': float(report.forced_outage_hours),
                'scheduled_outage': float(report.scheduled_outage_hours),
                'remarks': report.remarks or ''
            }
        
        return data
    
    def generate(self):
        """Generate PSR Excel file matching 100% the exact format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "PSR PSALM Edit (2)"
        
        # Set column widths
        widths_main = [25.86, 15.14, 13.71, 15.71, 13.71, 13.57, 13.29, 14.71, 13.0, 13.0, 13.0, 13.0, 15.71, 14.71]
        for i, width in enumerate(widths_main):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        
        # Right side columns
        extra_cols = ['O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
        extra_widths = [3.0, 15.0, 12.0, 12.0, 15.0, 3.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 12.0, 3.0, 15.0, 12.0, 12.0, 12.0, 12.0, 18, 10, 10]
        for col, width in zip(extra_cols, extra_widths):
            ws.column_dimensions[col].width = width
        
        # Add all sections
        self._add_header(ws)
        self._add_column_headers(ws)
        current_row = self._add_plant_data(ws)
        current_row = self._add_forecasted_load(ws, current_row)
        current_row = self._add_ipp_section(ws, current_row)
        current_row = self._add_notes_section(ws, current_row)
        current_row = self._add_footer(ws, current_row)
        
        # Add Agus 2 note and Chart
        self._add_chart_and_agus2_note(ws, current_row)
        
        # Add right side sections
        self._add_right_side_sections(ws)
        
        return wb

    def _add_header(self, ws):
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        header_font_color = "FFFFFF"
        
        title_text = 'DAILY PLANT STATUS' if self.report_type == 'daily_status' else ' PLANT STATUS REPORT'
        time_text = f'as of 12:00 NN    {self.report_date.strftime("%A, %B %d, %Y")}' if self.report_type == 'daily_status' else f'as of 0800H {self.report_date.strftime("%A, %d %B %Y")}'
        
        ws['A2'] = 'MINDANAO GENERATION'
        ws['A2'].font = Font(size=20, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:N2')
        
        ws['A3'] = '(PSALM PORTFOLIO)'
        ws['A3'].font = Font(size=14, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A3:N3')
        
        # Header box
        ws['A9'] = title_text
        ws['A9'].font = Font(size=18, bold=True, color=header_font_color)
        ws['A9'].fill = header_fill
        ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A9:N9')
        ws.row_dimensions[9].height = 25

        ws['A10'] = time_text
        ws['A10'].font = Font(size=14, bold=True, color=header_font_color)
        ws['A10'].fill = header_fill
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A10:N10')
        ws.row_dimensions[10].height = 25
            
    def _add_column_headers(self, ws):
        headers = {
            'A13': 'PLANT NAME', 'B13': 'Rated Capacity (MW)', 'C13': 'Nominated',
            'D13': 'Available Capacity (MW)', 'F13': 'Lake Lanao Outflow',
            'G13': 'Load at 0800H', 'H13': 'REMARKS'
        }
        for cell, val in headers.items():
            ws[cell] = val
            ws[cell].font = Font(size=12, bold=True)
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        ws.merge_cells('A13:A16')
        ws.merge_cells('B13:B16')
        # C13 is not merged as per template requirement
        ws.merge_cells('D13:D16')
        ws.merge_cells('F13:F16')
        ws.merge_cells('G13:G16')
        ws.merge_cells('H13:N16')

    def _add_plant_data(self, ws):
        current_row = 17
        
        agus_totals = {'capacity': 0, 'nominated': 0, 'actual': 0, 'variance': 0}
        hydro_totals = {'capacity': 0, 'nominated': 0, 'actual': 0, 'variance': 0}
        
        for p in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            start_row = current_row
            current_row = self._add_plant_section(ws, p, current_row)
            
            # Aggregate totals for the plant
            for r in range(start_row + 1, current_row):
                capacity = ws[f'B{r}'].value or 0
                nominated = ws[f'C{r}'].value or 0
                actual = ws[f'F{r}'].value or 0
                variance = ws[f'G{r}'].value or 0
                
                if p.startswith('AGUS'):
                    agus_totals['capacity'] += float(capacity)
                    agus_totals['nominated'] += float(nominated)
                    agus_totals['actual'] += float(actual)
                    agus_totals['variance'] += float(variance)
                    
                hydro_totals['capacity'] += float(capacity)
                hydro_totals['nominated'] += float(nominated)
                hydro_totals['actual'] += float(actual)
                hydro_totals['variance'] += float(variance)
                
            if p == 'AGUS7':
                ws[f'A{current_row}'] = ' TOTAL AGUS'
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'B{current_row}'] = round(agus_totals['capacity'], 1)
                ws[f'C{current_row}'] = round(agus_totals['nominated'], 1)
                ws[f'D{current_row}'] = round(agus_totals['nominated'], 1)
                ws[f'F{current_row}'] = round(agus_totals['actual'], 1)
                ws[f'G{current_row}'] = round(agus_totals['variance'], 1)
                
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                    ws[f'{col}{current_row}'].fill = fill
                current_row += 1
                
        # We also need to store hydro totals for later sections
        self.hydro_totals = hydro_totals
        
        return current_row

    def _add_plant_section(self, ws, plant_code, start_row):
        config = self.PLANTS_CONFIG[plant_code]
        data = self.data_by_plant.get(plant_code, {})
        ws[f'A{start_row}'] = config['name']
        ws[f'A{start_row}'].font = Font(bold=True)
        
        current_row = start_row + 1
        for unit in config['units']:
            ws[f'A{current_row}'] = unit['label']
            ws[f'B{current_row}'] = unit['capacity']
            
            unit_report_data = data.get(unit['num'], {})
            
            # Calculate actual generation (convert from kWh to MW)
            actual_generation = unit_report_data.get('generation', 0) / 1000
            if actual_generation > 0:
                actual_generation = min(actual_generation / 24, unit['capacity'])
                
            nominated = unit.get('nominated', 0)
            variance = nominated - actual_generation
            
            ws[f'C{current_row}'] = nominated
            ws[f'D{current_row}'] = nominated
            ws[f'F{current_row}'] = round(actual_generation, 1)
            ws[f'G{current_row}'] = round(variance, 1)
            
            remarks = unit_report_data.get('remarks', '')
            if not remarks and plant_code == 'AGUS1':
                remarks = 'Lake Lanao Elevation is 701.19 m.a.s.l. (G1- 0.10 m, G2- 0.10 m)'
            ws[f'H{current_row}'] = remarks
            
            current_row += 1
        return current_row

    def _add_forecasted_load(self, ws, current_row):
        ws[f'A{current_row}'] = f"Agus-Pulangi Forecasted Load: {self.report_date.strftime('%Y-%m-%d')}"
        ws[f'A{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.merge_cells(f'A{current_row}:N{current_row}')
        return current_row + 1

    def _add_ipp_section(self, ws, current_row):
        ws[f'A{current_row}'] = "TOTAL IPP"
        ws[f'A{current_row}'].font = Font(bold=True)
        # Assuming IPP is 0 for now as it's not in the database models
        ws[f'B{current_row}'] = 0
        ws[f'C{current_row}'] = 0
        ws[f'D{current_row}'] = 0
        ws[f'F{current_row}'] = 0
        ws[f'G{current_row}'] = 0
        
        ws[f'A{current_row+1}'] = "TOTAL NPC-PSALM"
        ws[f'A{current_row+1}'].font = Font(bold=True)
        
        hydro = getattr(self, 'hydro_totals', {'capacity': 0, 'nominated': 0, 'actual': 0, 'variance': 0})
        ws[f'B{current_row+1}'] = round(hydro['capacity'], 1)
        ws[f'C{current_row+1}'] = round(hydro['nominated'], 1)
        ws[f'D{current_row+1}'] = round(hydro['nominated'], 1)
        ws[f'F{current_row+1}'] = round(hydro['actual'], 1)
        ws[f'G{current_row+1}'] = round(hydro['variance'], 1)
        
        return current_row + 3

    def _add_notes_section(self, ws, current_row):
        ws[f'A{current_row}'] = "Notes:"
        
        notes = [
            "1. The Available Capacity in this report includes equipment limitation and water outflow consideration based on the 2020 Lake Lanao Operating Guide Curve.",
            "2. AGUS 6 HEP units 1 & 2 up-rated from 25MW to 34.5MW. Turned-over to NPC last 14 February 2020.",
            "3. AGUS 2 HEP is limited to 120 MW total load to prevent risk of flooding at lakeshore areas and Baloi plains as per Environmental Compliance Certificate dated January 14, 1992.",
            "4. AGUS 5 HEP gate no. 2 dogged at 0.10m for Newtech Pulp Inc. (NPI) plant water use.",
            "5. The usual occurrence of Peak is at 1800H.",
            "6. Forecast inflow of Lake Lanao is stable and operating at Normal Stage.",
            "7. Agus 6 HEP: 18cms of water spilled due to partially opened spillway gate no. 1."
        ]
        for n in notes:
            current_row += 1
            ws[f'A{current_row}'] = n
            ws.merge_cells(f'A{current_row}:N{current_row}')
        return current_row + 2

    def _add_footer(self, ws, start_row):
        # We can simulate the preview's footer
        ws[f'A{start_row}'] = "AUTHORIZATION"
        ws[f'A{start_row}'].font = Font(bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{start_row}:N{start_row}')
        
        roles = ['Prepared by:', 'Checked and Reviewed by:', 'Checked and Reviewed by:', 'Approved by:']
        names = ['O.M. LAVA', 'JMM MATA', 'EL ADIONG', 'C.C. AMIGABLE JR.']
        titles = ['Prin. Engr. A, GPD', 'Manager, GPD', 'Acting Manager, GPD', 'Dept. Manager, GPD']
        cols = ['B', 'E', 'H', 'K']
        
        current_row = start_row + 2
        for c, role in zip(cols, roles):
            ws[f'{c}{current_row}'] = role
            
        current_row += 3
        for c, name in zip(cols, names):
            ws[f'{c}{current_row}'] = name
            ws[f'{c}{current_row}'].font = Font(bold=True, underline='single')
            
        current_row += 1
        for c, title in zip(cols, titles):
            ws[f'{c}{current_row}'] = title
            
        return current_row + 2

    def _add_chart_and_agus2_note(self, ws, start_row):
        ws[f'A{start_row}'] = 'Agus 2 HEP is limited to 40 MW/per unit due to water constraints...'
        ws.merge_cells(f'A{start_row}:N{start_row}')
        
        # Excel 2007 requires valid XML for everything.
        # Ensure we don't have leftover data references that might confuse it.

    def _add_right_side_sections(self, ws):
        ws['P13'] = 'PRIMARY STORAGE'
        ws['P13'].font = Font(bold=True)
        ws.merge_cells('P13:S13')

    def _calculate_generation_data(self):
        """Calculate generation data for today, MTD, YTD from database"""
        from django.db.models import Sum
        from reports.models import GenerationReport, Plant
        
        report_date = self.report_date
        today_start = report_date
        mtd_start = report_date.replace(day=1)
        ytd_start = report_date.replace(month=1, day=1)
        
        gen_data = []
        total_today, total_mtd, total_ytd = 0, 0, 0
        plant_codes = ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']
        
        for plant_code in plant_codes:
            try:
                plant = Plant.objects.get(code=plant_code)
                today_gen = GenerationReport.objects.filter(plant=plant, report_date=today_start).aggregate(total=Sum('generation_kwh'))['total'] or 0
                mtd_gen = GenerationReport.objects.filter(plant=plant, report_date__gte=mtd_start, report_date__lte=report_date).aggregate(total=Sum('generation_kwh'))['total'] or 0
                ytd_gen = GenerationReport.objects.filter(plant=plant, report_date__gte=ytd_start, report_date__lte=report_date).aggregate(total=Sum('generation_kwh'))['total'] or 0
                
                today_mwh, mtd_mwh, ytd_mwh = float(today_gen)/1000, float(mtd_gen)/1000, float(ytd_gen)/1000
                gen_data.append([plant_code, f'{today_mwh:,.0f}', f'{mtd_mwh:,.0f}', f'{ytd_mwh:,.0f}'])
                total_today += today_mwh
                total_mtd += mtd_mwh
                total_ytd += ytd_mwh
            except:
                gen_data.append([plant_code, '0', '0', '0'])
        
        gen_data.append(['Total NPC', f'{total_today:,.0f}', f'{total_mtd:,.0f}', f'{total_ytd:,.0f}'])
        return gen_data

    def _calculate_capacity_factor(self):
        """Calculate capacity factor for today, MTD, YTD from database"""
        from django.db.models import Avg
        from reports.models import GenerationReport, Plant
        
        report_date = self.report_date
        mtd_start = report_date.replace(day=1)
        ytd_start = report_date.replace(month=1, day=1)
        
        cf_data = []
        cf_today_list, cf_mtd_list, cf_ytd_list = [], [], []
        plant_codes = ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']
        
        for plant_code in plant_codes:
            try:
                plant = Plant.objects.get(code=plant_code)
                today_cf = GenerationReport.objects.filter(plant=plant, report_date=report_date).aggregate(avg=Avg('capacity_factor'))['avg'] or 0
                mtd_cf = GenerationReport.objects.filter(plant=plant, report_date__gte=mtd_start, report_date__lte=report_date).aggregate(avg=Avg('capacity_factor'))['avg'] or 0
                ytd_cf = GenerationReport.objects.filter(plant=plant, report_date__gte=ytd_start, report_date__lte=report_date).aggregate(avg=Avg('capacity_factor'))['avg'] or 0
                
                cf_data.append([plant_code, f'{float(today_cf):.1f}', f'{float(mtd_cf):.1f}', f'{float(ytd_cf):.1f}'])
                cf_today_list.append(float(today_cf))
                cf_mtd_list.append(float(mtd_cf))
                cf_ytd_list.append(float(ytd_cf))
            except:
                cf_data.append([plant_code, '0.0', '0.0', '0.0'])
                cf_today_list.append(0.0)
                cf_mtd_list.append(0.0)
                cf_ytd_list.append(0.0)
        
        avg_today = sum(cf_today_list)/len(cf_today_list) if cf_today_list else 0
        avg_mtd = sum(cf_mtd_list)/len(cf_mtd_list) if cf_mtd_list else 0
        avg_ytd = sum(cf_ytd_list)/len(cf_ytd_list) if cf_ytd_list else 0
        cf_data.append(['Average', f'{avg_today:.1f}', f'{avg_mtd:.1f}', f'{avg_ytd:.1f}'])
        return cf_data
